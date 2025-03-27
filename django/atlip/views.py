# atlip/views.py
from .models import GroupPortfolio, PatentPortfolio, BrandPortfolio , OwnerPortfolio , AgentPortfolio, DomainPortfolio, ModelPortfolio
from .serializers import GroupPortfolioSerializer , PatentPortfolioSerializer, BrandPortfolioSerializer, OwnerPortfolioSerializer, AgentPortfolioSerializer, DomainPortfolioSerializer, ModelPortfolioSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated  # Require authentication
from rest_framework.parsers import MultiPartParser, FormParser
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from rest_framework.decorators import action
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
from django.utils import timezone
from rest_framework.views import APIView
from django.core.files.uploadedfile import InMemoryUploadedFile
from .filters import GroupPortfolioFilter , BrandPortfolioFilter , OwnerPortfolioFilter, AgentPortfolioFilter, DomainPortfolioFilter, ModelPortfolioFilter


from pathlib import Path
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import sys
import os
from datetime import datetime
import pandas as pd 
import logging
from urllib.parse import urljoin
from PIL import Image
from io import BytesIO
import requests
import xlsxwriter


# Configure logging
logger = logging.getLogger(__name__)


class GroupPortfolioViewSet(viewsets.ModelViewSet):
    queryset = GroupPortfolio.objects.all()
    serializer_class = GroupPortfolioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]

    filterset_class = GroupPortfolioFilter  # Use the custom filter class
    search_fields = ['group_name', 'group_email', 'group_phone', 'group_address', 'group_vat', 'group_contact_person']

    def list(self, request, *args, **kwargs):
        """Fetch all Brand records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Brand record by ID."""
        try:
            brand = GroupPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(brand)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except GroupPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new brand record."""
        print( request.data , "post request")
        # Get the uploaded file
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print( request.data)
        try:
            brand = GroupPortfolio.objects.get(pk=pk)
        except GroupPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(brand, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single brand record by ID."""
        try:
            brand = GroupPortfolio.objects.get(pk=pk)
        except GroupPortfolio.DoesNotExist:
            return Response({"error": "brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( brand, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single brand record by ID."""
        try:
            brand = GroupPortfolio.objects.get(pk=pk)
            brand.delete()
            return Response({"message": "Brand deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except GroupPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple brand records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = GroupPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} brands deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = GroupPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "Group Name", "Group Email", "Group Address", "Group Phone", 
            "Group VAT", "Group Contact Person"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.group_name, row.group_email, row.group_address, row.group_phone, row.group_vat, row.group_contact_person
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else 'N/A', text_wrap_format)

            # Handle Image Download & Insertion with formatting
            
            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_Groups.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple brand records by IDs including logo.
        """
        print("Received Data:", request.data)
        # print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        # logo = request.FILES.get('logo')

        # print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data :
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for group in GroupPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr( group, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr( group, key, value)

            
            if changes:
                print(f"Updating Brand ID { group.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            group_data_before_save = model_to_dict( group)
            print(f"Brand Data Before Save (ID: {group.id}):", group_data_before_save)
            group.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} brands updated successfully"}, status=status.HTTP_200_OK)



class UploadXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 :
            # Read XLS file
            df = pd.read_excel(file_path, engine="openpyxl")  # Use openpyxl for XLSX
            print(df.head())
            print(df.columns)
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            # Print renamed columns
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = ["group_name", "group_email", "group_phone", "group_address", "group_vat", "group_contact_person"]
            missing_columns = set(required_columns) - set(df.columns)
            print( missing_columns , 'missing_columns' )
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Iterate through rows and save data
            records = []
            print( "starting iterations")
            for _, row in df.iterrows():
                print(row)
                data = {
                    "group_name": row["group_name"],
                    "group_email": row["group_email"],
                    "group_phone": row["group_phone"],
                    "group_address": row["group_address"],
                    "group_vat": row["group_vat"],
                    "group_contact_person": row["group_contact_person"],
                }
                serializer = GroupPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        #except Exception as e:
        #    return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        #finally:
        #    # Clean up temporary file
        #    if os.path.exists(file_path):
        #        os.remove(file_path)



class GroupPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = ["group_name", "group_email", "group_phone", "group_address", "group_vat", "group_contact_person"]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="group_portfolio_template.xlsx"'
        
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="GroupPortfolio", index=False)

        return response
    
# Views in Patent Portfolio 

class PatentPortfolioViewSet(viewsets.ModelViewSet):
    queryset = PatentPortfolio.objects.all().order_by('-updated_date')
    serializer_class = PatentPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = [
        'patent_title', 'territory', 'p_type','owner_name', 'agent_name','group_name', 'next_annuity', 'status'
    ]
    search_fields = [
        'patent_title', 'territory', 'p_type','owner_name', 'agent_name','group_name', 'next_annuity', 'status'
    ]

    def list(self, request, *args, **kwargs):
        """Fetch all patent records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single patent record by ID."""
        try:
            patent = PatentPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(patent)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except PatentPortfolio.DoesNotExist:
            return Response({"error": "Patent not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new patent record."""
        print( request.data , "post request")
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        try:
            patent = PatentPortfolio.objects.get(pk=pk)
        except PatentPortfolio.DoesNotExist:
            return Response({"error": "Patent not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(patent, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single patent record by ID."""
        try:
            patent = PatentPortfolio.objects.get(pk=pk)
        except PatentPortfolio.DoesNotExist:
            return Response({"error": "Patent not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(patent, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single patent record by ID."""
        try:
            patent = PatentPortfolio.objects.get(pk=pk)
            patent.delete()
            return Response({"message": "Patent deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except PatentPortfolio.DoesNotExist:
            return Response({"error": "Patent not found"}, status=status.HTTP_404_NOT_FOUND)

    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple brand records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = PatentPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} Patent deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = PatentPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "patent_title",
            "territory",
            "p_type",
            "priority",
            "application_date",
            "application_no",
            "publication_date",
            "publication_no",
            "registration_date",
            "registration_no",
            "next_annuity",
            "annuity_no",
            "owner_name",
            "owner_email",  
            "owner_phone",
            "owner_address",
            "owner_vat",
            "owner_contact_person",
            "agent_name",
            "agent_email",
            "agent_phone",
            "agent_address",
            "agent_vat",
            "agent_contact_person",
            "group_name",
            "group_email",
            "group_address",
            "group_phone",
            "group_vat",
            "group_contact_person",
            "inventor",
            "status",
            "comments"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.patent_title,
                row.territory,
                row.p_type,
                row.priority,
                row.application_date,
                row.application_no,
                row.publication_date,
                row.publication_no,
                row.registration_date,
                row.registration_no,
                row.next_annuity,
                row.annuity_no,
                row.owner_name,
                row.owner_email,  
                row.owner_phone,
                row.owner_address,
                row.owner_vat,
                row.owner_contact_person,
                row.agent_name,
                row.agent_email,
                row.agent_phone,
                row.agent_address,
                row.agent_vat,
                row.agent_contact_person,
                row.group_name,
                row.group_email,
                row.group_address,
                row.group_phone,
                row.group_vat,
                row.group_contact_person,
                row.inventor,
                row.status,
                row.comments

                
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else '', text_wrap_format)

            # Handle Image Download & Insertion with formatting
            
            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_Patents.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple brand records by IDs including logo.
        """
        print("Received Data:", request.data)
        # print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        # logo = request.FILES.get('logo')

        # print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data :
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for patent in PatentPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr( patent, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr( patent, key, value)

            
            if changes:
                print(f"Updating Brand ID { patent.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            patent_data_before_save = model_to_dict( patent )
            print(f"Brand Data Before Save (ID: { patent.id}):", patent_data_before_save)
            patent.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} brands updated successfully"}, status=status.HTTP_200_OK)
      


class UploadPatentXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        try:
            # Read XLS file
            df = pd.read_excel(file_path, engine="openpyxl")  # Use openpyxl for XLSX
            print(df.head())
            print(df.columns)
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

             # Convert date columns to proper date format
            date_columns = ["application_date", "publication_date", "registration_date", "next_annuity"]
            for col in date_columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

            # Print renamed columns
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                "patent_title",
            "territory",
            "p_type",
            "priority",
            "application_date",
            "application_no",
            "publication_date",
            "publication_no",
            "registration_date",
            "registration_no",
            "next_annuity",
            "annuity_no",
            "owner_name",
            "owner_email",  
            "owner_phone",
            "owner_address",
            "owner_vat",
            "owner_contact_person",
            "agent_name",
            "agent_email",
            "agent_phone",
            "agent_address",
            "agent_vat",
            "agent_contact_person",
            "group_name",
            "group_email",
            "group_address",
            "group_phone",
            "group_vat",
            "group_contact_person",
            "inventor",
            "status",
            "comments"
            ]
            missing_columns = set(required_columns) - set(df.columns)
            print(missing_columns, "missing_columns")
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Iterate through rows and save data
            records = []
            print("Starting iterations")
            for _, row in df.iterrows():
                print(row)
                data = {
                    "patent_title" : row["patent_title"],
                    "territory" : row["territory"],
                    "p_type" : row["p_type"],
                    "priority" : row["priority"],
                    "application_date" : row["application_date"],
                    "application_no" : row["application_no"],
                    "publication_date" : row["publication_date"],
                    "publication_no" : row["publication_no"],
                    "registration_date" : row["registration_date"],
                    "registration_no" : row["registration_no"],
                    "next_annuity" : row["next_annuity"],
                    "annuity_no" : row["annuity_no"],
                    "owner_name" : row["owner_name"],
                    "owner_email" : row["owner_email"],  
                    "owner_phone" : row["owner_phone"],
                    "owner_address" : row["owner_address"],
                    "owner_vat" : row["owner_vat"],
                    "owner_contact_person" : row["owner_contact_person"],
                    "agent_name" : row["agent_name"],
                    "agent_email" : row["agent_email"],
                    "agent_phone" : row["agent_phone"],
                    "agent_address" : row["agent_address"],
                    "agent_vat" : row["agent_vat"],
                    "agent_contact_person" : row["agent_contact_person"] ,
                    "group_name" : row["group_name"],
                    "group_email" : row["group_email"],
                    "group_address" : row["group_address"],
                    "group_phone" : row["group_phone"],
                    "group_vat" : row["group_vat"],
                    "group_contact_person" : row["group_contact_person"],
                    "inventor" : row["inventor"],
                    "status" : row["status"],
                    "comments" : row["comments"]

                    
                }
                serializer = PatentPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print(serializer.errors)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finally:
            # Clean up temporary file
            if os.path.exists(file_path):
                os.remove(file_path)


class PatentPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            "patent_title",
            "territory",
            "p_type",
            "priority",
            "application_date",
            "application_no",
            "publication_date",
            "publication_no",
            "registration_date",
            "registration_no",
            "next_annuity",
            "annuity_no",
            "owner_name",
            "owner_email",  
            "owner_phone",
            "owner_address",
            "owner_vat",
            "owner_contact_person",
            "agent_name",
            "agent_email",
            "agent_phone",
            "agent_address",
            "agent_vat",
            "agent_contact_person",
            "group_name",
            "group_email",
            "group_address",
            "group_phone",
            "group_vat",
            "group_contact_person",
            "inventor",
            "status",
            "comments"
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="patent_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="PatentPortfolio", index=False)

        return response
    

# class UploadLogoView(APIView):
#     parser_classes = (MultiPartParser, FormParser)
#     permission_classes = [IsAuthenticated]  # Ensures only authenticated users can upload

#     def post(self, request, *args, **kwargs):
#         try:
#             # Log the headers
#             logger.info("Received headers: %s", request.headers)
#             print("Received headers:", request.headers)

#             # Log form data
#             logger.info("Received form data: %s", request.data)
#             print("Received form data:", request.data)

#             # Log uploaded files
#             logger.info("Received files: %s", request.FILES)
#             print("Received files:", request.FILES)

#             logo_file = request.FILES.get("logo")
#             if not logo_file:
#                 return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#             # Generate dynamic upload path based on date
#             today = datetime.today()
#             upload_dir = f"logo/{today.year}/{today.month:02d}/{today.day:02d}/"

#             # Save the file
#             file_path = os.path.join(upload_dir, logo_file.name)
#             saved_path = default_storage.save(file_path, ContentFile(logo_file.read()))
#             full_file_url = urljoin(settings.MEDIA_URL, saved_path)

#             logger.info("File saved at: %s", full_file_url)
#             print("File saved at:", full_file_url)

#             return Response(
#                 {"logo": full_file_url},
#                 status=status.HTTP_201_CREATED
#             )

#         except Exception as e:
#             logger.error("File upload failed: %s", str(e))
#             print("File upload failed:", str(e))
#             return Response(
#                 {"error": f"File upload failed: {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )


# Views in brand/trademark Portfolio 

class BrandPortfolioViewSet(viewsets.ModelViewSet):
    queryset = BrandPortfolio.objects.all().order_by('-updated_date')
    serializer_class = BrandPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  BrandPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ["title",  "territory", "b_Type" ,  "deadline", "status",
        "owner_name", "agent_name" , "group_name", "affidavit"]


    def list(self, request, *args, **kwargs):
        """Fetch all Brand records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(brand)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new brand record."""
        print( request.data , "post request")
        # Get the uploaded file
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print( request.data)
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(brand, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( brand, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
            brand.delete()
            return Response({"message": "Brand deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple brand records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = BrandPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} brands deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = BrandPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "Title", "Territory", "Type", "Application Date", "Application No", "Registration Date", "Registration No", 
            "Classes", "Deadline", "Affidavit", "Owner Name", "Owner Email", "Owner Phone", "Owner Address", 
            "Owner VAT", "Owner Contact Person", "Agent Name", "Agent Email", "Agent Phone", "Agent Address", 
            "Agent VAT", "Agent Contact Person", "Group Name", "Group Email", "Group Address", "Group Phone", 
            "Group VAT", "Group Contact Person", "Comments", "Status", "Updated Date", "Logo"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.title, row.territory, row.b_Type, str(row.application_date), row.application_no,
                str(row.registration_date), row.registration_no, row.classes, str(row.deadline), str(row.affidavit),
                row.owner_name, row.owner_email, row.owner_phone, row.owner_address, row.owner_vat, row.owner_contact_person,
                row.agent_name, row.agent_email, row.agent_phone, row.agent_address, row.agent_vat, row.agent_contact_person,
                row.group_name, row.group_email, row.group_address, row.group_phone, row.group_vat, row.group_contact_person,
                row.comments, row.status, str(row.updated_date)
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else 'N/A', text_wrap_format)

            # Handle Image Download & Insertion with formatting
            if row.logo and row.logo != 'na.png':
                logo_url = request.build_absolute_uri(f'{settings.MEDIA_URL}{row.logo}')
                try:
                    image_response = requests.get(logo_url, stream=True)
                    if image_response.status_code == 200:
                        image_path = f"/tmp/{row.id}.png"
                        with open(image_path, "wb") as img_file:
                            img_file.write(image_response.content)

                        from PIL import Image
                        img = Image.open(image_path)
                        width, height = img.size
                        scale_x = 100 / width
                        scale_y = 100 / height
                        scale = min(scale_x, scale_y)
                        worksheet.set_row(row_num, 100)  # Set row height to a fixed size
                        # Write a blank cell with border format before inserting the image
                        worksheet.write_blank(row_num, len(headers) - 1, '', border_format)

                        # Insert image after ensuring the cell has a border
                        worksheet.insert_image(row_num, len(headers) - 1, image_path, {'x_scale': scale, 'y_scale': scale})
                    else:
                        worksheet.write(row_num, len(headers) - 1, "Image not available", border_format)
                except Exception as e:
                    print(f"Failed to download image for {row.title}: {e}")
                    worksheet.write(row_num, len(headers) - 1, "Error loading image", border_format)
            else:
                worksheet.write(row_num, len(headers) - 1, "No Image", border_format)

            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_brands.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple brand records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for brand in BrandPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(brand, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr(brand, key, value)

            # Check and update logo only if provided
            print("before" , logo)
            if logo is not None:
                print("after" , logo)
                current_logo = brand.logo.url if brand.logo else None
                changes['logo'] = {"old": current_logo, "new": logo.name}
                brand.logo = logo

            if changes:
                print(f"Updating Brand ID {brand.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            brand_data_before_save = model_to_dict(brand)
            print(f"Brand Data Before Save (ID: {brand.id}):", brand_data_before_save)
            brand.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} brands updated successfully"}, status=status.HTTP_200_OK)



class UploadBrandXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            image_loader = SheetImageLoader(sheet)  # Load images

            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            df["application_date"] = df["application_date"].dt.strftime("%Y-%m-%d")
            df["registration_date"] = df["registration_date"].dt.strftime("%Y-%m-%d")
            df["deadline"] = df["deadline"].dt.strftime("%Y-%m-%d")
            df["affidavit"] = df["affidavit"].dt.strftime("%Y-%m-%d")
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                "title", "logo", "b_type", "territory", "application_date", "application_no",
                "registration_date", "registration_no", "classes", "deadline", "affidavit",
                "owner_name", "owner_email", "owner_phone", "owner_address", "owner_vat",
                "owner_contact_person", "agent_name", "agent_email", "agent_phone",
                "agent_address", "agent_vat", "agent_contact_person", "group_name",
                "group_email", "group_address", "group_phone", "group_vat",
                "group_contact_person", "status", "comments"
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Get today's date for folder structuring
            today = datetime.date.today()
            year, month, day = today.year, today.month, today.day
            image_folder = Path(settings.MEDIA_ROOT) / "logos" / str(year) / str(month) / str(day)
            image_folder.mkdir(parents=True, exist_ok=True)  # Create if not exists

            # Process rows
            records = []
            max_id = BrandPortfolio.objects.all().count()  # Get max ID for unique naming

            print("Starting iterations")
            for index, row in df.iterrows():
                print( index , row )
                logo_image_path = None

                # Extract and save image if present
                image_cell = f"B{index + 2}"  # Assuming logos are in column B
                if image_loader.image_in(image_cell):
                    print("Saving image from cell")
                    img = image_loader.get(image_cell)
        
                    max_id += 1
                    image_name = f"image_{max_id}.png"
                    image_path = image_folder / image_name
                    img.save(image_path)
                    logo_image_path = f"logos/{year}/{month}/{day}/{image_name}"
                    print("Image loaded")

                    # Convert the image to an in-memory file for posting
                    image_io = BytesIO()
                    img.save(image_io, format="PNG")
                    image_io.seek(0)  # Reset pointer to start

                    # Convert BytesIO to InMemoryUploadedFile
                    logo_image_file = InMemoryUploadedFile(
                        file=image_io,  # The in-memory file
                        field_name=None,
                        name=image_name,  # File name
                        content_type="image/png",
                        size=sys.getsizeof(image_io),
                        charset=None
                    )

                # Prepare data for saving
                # Format Date type fields \
                print( type( row["application_date"]) , type(logo_image_file) , "type check")

                data = {
                    "title": row["title"],
                    "logo": logo_image_file ,  
                    "b_Type": row["b_type"],
                    "territory": row["territory"],
                    "application_date": row["application_date"],
                    "application_no": row["application_no"],
                    "registration_date": row["registration_date"],
                    "registration_no": row["registration_no"],
                    "classes": row["classes"],
                    "deadline": row["deadline"],
                    "affidavit": row["affidavit"],
                    "owner_name": row["owner_name"],
                    "owner_email": row["owner_email"],
                    "owner_phone": row["owner_phone"],
                    "owner_address": row["owner_address"],
                    "owner_vat": row["owner_vat"],
                    "owner_contact_person": row["owner_contact_person"],
                    "agent_name": row["agent_name"],
                    "agent_email": row["agent_email"],
                    "agent_phone": row["agent_phone"],
                    "agent_address": row["agent_address"],
                    "agent_vat": row["agent_vat"],
                    "agent_contact_person": row["agent_contact_person"],
                    "group_name": row["group_name"],
                    "group_email": row["group_email"],
                    "group_address": row["group_address"],
                    "group_phone": row["group_phone"],
                    "group_vat": row["group_vat"],
                    "group_contact_person": row["group_contact_person"],
                    "status": row["status"],
                    "comments": row["comments"],
                }

                serializer = BrandPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class BrandPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            "title", 
                "logo",
                "b_Type", 
                "territory",
                "application_date",
                "application_no",
                "registration_date",
                "registration_no",
                "classes",
                "deadline",
                "affidavit",
                "owner_name", 
                "owner_email", 
                "owner_phone", 
                "owner_address", 
                "owner_vat", 
                "owner_contact_person", 
                "agent_name", 
                "agent_email" , 
                "agent_phone", 
                "agent_address", 
                "agent_vat", 
                "agent_contact_person",
                "group_name",
                "group_email",
                "group_address",
                "group_phone",
                "group_vat",
                "group_contact_person",
                "status",
                "comments"
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="patent_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="PatentPortfolio", index=False)

        return response
    


# Views in Owner Portfolio 

class OwnerPortfolioViewSet(viewsets.ModelViewSet):
    queryset = OwnerPortfolio.objects.all().order_by('-updated_date')
    serializer_class = OwnerPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  OwnerPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ['owner_name', 'owner_email', 'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person']


    def list(self, request, *args, **kwargs):
        """Fetch all Owner records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Owner record by ID."""
        try:
            Owner = OwnerPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(Owner)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except OwnerPortfolio.DoesNotExist:
            return Response({"error": "Owner not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new Owner record."""
        print( request.data , "post request")
        # Get the uploaded file
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print( request.data)
        try:
            Owner = OwnerPortfolio.objects.get(pk=pk)
        except OwnerPortfolio.DoesNotExist:
            return Response({"error": "Owner not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(Owner, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single Owner record by ID."""
        try:
            Owner = OwnerPortfolio.objects.get(pk=pk)
        except OwnerPortfolio.DoesNotExist:
            return Response({"error": "Owner not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( Owner, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single Owner record by ID."""
        try:
            Owner = OwnerPortfolio.objects.get(pk=pk)
            Owner.delete()
            return Response({"message": "Owner deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except OwnerPortfolio.DoesNotExist:
            return Response({"error": "Owner not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple Owner records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = OwnerPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} Owners deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = OwnerPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "Owner Name", "Owner Email", "Owner Phone", "Owner Address", 
            "Owner VAT", "Owner Contact Person"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.owner_name, row.owner_email, row.owner_phone, row.owner_address, row.owner_vat, row.owner_contact_person
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else 'N/A', text_wrap_format)

            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_Owners.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple Owner records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for Owner in OwnerPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(Owner, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr(Owner, key, value)

            # Check and update logo only if provided
            print("before" , logo)
            if logo is not None:
                print("after" , logo)
                current_logo = Owner.logo.url if Owner.logo else None
                changes['logo'] = {"old": current_logo, "new": logo.name}
                Owner.logo = logo

            if changes:
                print(f"Updating Owner ID {Owner.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            Owner_data_before_save = model_to_dict(Owner)
            print(f"Owner Data Before Save (ID: {Owner.id}):", Owner_data_before_save)
            Owner.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} Owners updated successfully"}, status=status.HTTP_200_OK)



class UploadOwnerXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            
            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                "owner_name", "owner_email", "owner_phone", "owner_address", "owner_vat",
                "owner_contact_person"
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Get today's date for folder structuring
            from datetime import datetime
            today = datetime.today()
            year, month, day = today.year, today.month, today.day
            image_folder = Path(settings.MEDIA_ROOT) / "logos" / str(year) / str(month) / str(day)
            image_folder.mkdir(parents=True, exist_ok=True)  # Create if not exists

            # Process rows
            records = []
            max_id = OwnerPortfolio.objects.all().count()  # Get max ID for unique naming

            print("Starting iterations")
            for index, row in df.iterrows():
                print( index , row )
                data = {
                    "owner_name": row["owner_name"],
                    "owner_email": row["owner_email"],
                    "owner_phone": row["owner_phone"],
                    "owner_address": row["owner_address"],
                    "owner_vat": row["owner_vat"],
                    "owner_contact_person": row["owner_contact_person"],
                    
                }

                serializer = OwnerPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class OwnerPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            "owner_name", 
            "owner_email", 
            "owner_phone", 
            "owner_address", 
            "owner_vat", 
            "owner_contact_person", 
            
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="owner_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="OwnerPortfolio", index=False)

        return response
    

# Views in agent PortFolio 

# Views in Agent Portfolio 

class AgentPortfolioViewSet(viewsets.ModelViewSet):
    queryset = AgentPortfolio.objects.all().order_by('-updated_date')
    serializer_class = AgentPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  AgentPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ['Agent_name', 'Agent_email', 'Agent_phone', 'Agent_address', 'Agent_vat', 'Agent_contact_person']


    def list(self, request, *args, **kwargs):
        """Fetch all Agent records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Agent record by ID."""
        try:
            Agent = AgentPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(Agent)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except AgentPortfolio.DoesNotExist:
            return Response({"error": "Agent not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new Agent record."""
        print( request.data , "post request")
        # Get the uploaded file
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print( request.data)
        try:
            Agent = AgentPortfolio.objects.get(pk=pk)
        except AgentPortfolio.DoesNotExist:
            return Response({"error": "Agent not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(Agent, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single Agent record by ID."""
        try:
            Agent = AgentPortfolio.objects.get(pk=pk)
        except AgentPortfolio.DoesNotExist:
            return Response({"error": "Agent not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( Agent, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single Agent record by ID."""
        try:
            Agent = AgentPortfolio.objects.get(pk=pk)
            Agent.delete()
            return Response({"message": "Agent deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except AgentPortfolio.DoesNotExist:
            return Response({"error": "Agent not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple Agent records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = AgentPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} Agents deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = AgentPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "Agent Name", "Agent Email", "Agent Phone", "Agent Address", 
            "Agent VAT", "Agent Contact Person"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.agent_name, row.agent_email, row.agent_phone, row.agent_address, row.agent_vat, row.agent_contact_person,
                ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else 'N/A', text_wrap_format)

            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_Agents.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple Agent records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for Agent in AgentPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(Agent, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr(Agent, key, value)

            if changes:
                print(f"Updating Agent ID {Agent.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            Agent_data_before_save = model_to_dict(Agent)
            print(f"Agent Data Before Save (ID: {Agent.id}):", Agent_data_before_save)
            Agent.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} Agents updated successfully"}, status=status.HTTP_200_OK)



class UploadAgentXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            
            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                "agent_name", "agent_email", "agent_phone", "agent_address", "agent_vat",
                "agent_contact_person"
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Get today's date for folder structuring
            from datetime import datetime
            today = datetime.today()
            year, month, day = today.year, today.month, today.day
            
            # Process rows
            records = []
            max_id = AgentPortfolio.objects.all().count()  # Get max ID for unique naming

            print("Starting iterations")
            for index, row in df.iterrows():
                print( index , row )
                data = {
                    "agent_name": row["agent_name"],
                    "agent_email": row["agent_email"],
                    "agent_phone": row["agent_phone"],
                    "agent_address": row["agent_address"],
                    "agent_vat": row["agent_vat"],
                    "agent_contact_person": row["agent_contact_person"],
                    
                }

                serializer = AgentPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class AgentPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            "Agent_name", 
            "Agent_email", 
            "Agent_phone", 
            "Agent_address", 
            "Agent_vat", 
            "Agent_contact_person", 
            
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Agent_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="AgentPortfolio", index=False)

        return response
    

# Views of Domain 

# Views in domain/trademark Portfolio 

class DomainPortfolioViewSet(viewsets.ModelViewSet):
    queryset = DomainPortfolio.objects.all().order_by('-updated_date')
    serializer_class = DomainPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  DomainPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ["domain_name",  "extension",   "deadline", "status",
        "owner_name", "agent_name" , "group_name"]


    def list(self, request, *args, **kwargs):
        """Fetch all Domain records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Domain record by ID."""
        try:
            domain = DomainPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(domain)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except DomainPortfolio.DoesNotExist:
            return Response({"error": "Domain not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new domain record."""
        print( request.data , "post request")
        # Get the uploaded file
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print( request.data)
        try:
            domain = DomainPortfolio.objects.get(pk=pk)
        except DomainPortfolio.DoesNotExist:
            return Response({"error": "Domain not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(domain, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single domain record by ID."""
        try:
            domain = DomainPortfolio.objects.get(pk=pk)
        except DomainPortfolio.DoesNotExist:
            return Response({"error": "domain not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( domain, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single domain record by ID."""
        try:
            domain = DomainPortfolio.objects.get(pk=pk)
            domain.delete()
            return Response({"message": "Domain deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except DomainPortfolio.DoesNotExist:
            return Response({"error": "Domain not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple domain records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = DomainPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} domains deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = DomainPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
                    'domain_name', 'extension', 'creation_date', 'deadline', 
                    'owner_name', 'owner_email', 'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
                    'agent_name', 'agent_email', 'agent_phone', 'agent_address', 'agent_vat', 'agent_contact_person', 
                    'group_name', 'group_email', 'group_address', 'group_phone', 'group_vat', 'group_contact_person', 
                    'status', 'comments' 
                ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                    row.domain_name , row.extension , str(row.creation_date), str(row.deadline),  
                    row.owner_name, row.owner_email, row.owner_phone,  row.owner_address, row.owner_vat,  row.owner_contact_person, 
                    row.agent_name,  row.agent_email,  row.agent_phone,  row.agent_address , row.agent_vat,  row.agent_contact_person,  
                    row.group_name,  row.group_email,  row.group_address,  row.group_phone, row.group_vat,  row.group_contact_person,  
                    row.status,  row.comments 

            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else '', text_wrap_format)
                row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_domains.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple domain records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for domain in DomainPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(domain, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr(domain, key, value)

            

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            domain_data_before_save = model_to_dict(domain)
            print(f"Domain Data Before Save (ID: {domain.id}):", domain_data_before_save)
            domain.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} domains updated successfully"}, status=status.HTTP_200_OK)



class UploadDomainXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            image_loader = SheetImageLoader(sheet)  # Load images

            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            df["creation_date"] = df["creation_date"].dt.strftime("%Y-%m-%d")
            df["deadline"] = df["deadline"].dt.strftime("%Y-%m-%d")
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                'domain_name', 'extension', 'creation_date', 'deadline', 
                'owner_name', 'owner_email', 'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
                'agent_name', 'agent_email', 'agent_phone', 'agent_address', 'agent_vat', 'agent_contact_person', 
                'group_name', 'group_email', 'group_address', 'group_phone', 'group_vat', 'group_contact_person', 
                'status', 'comments'
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            
            # Process rows
            records = []
            max_id = DomainPortfolio.objects.all().count()  # Get max ID for unique naming

            print("Starting iterations")
            for index, row in df.iterrows():
                print( index , row )
                logo_image_path = None

                # Prepare data for saving
                # Format Date type fields \
                

                data = {
                    'domain_name' : row['domain_name'] , 
                    'extension' : row['extension'], 
                    'creation_date' : row['creation_date'], 
                    'deadline' : row['deadline'],
                    "owner_name": row["owner_name"],
                    "owner_email": row["owner_email"],
                    "owner_phone": row["owner_phone"],
                    "owner_address": row["owner_address"],
                    "owner_vat": row["owner_vat"],
                    "owner_contact_person": row["owner_contact_person"],
                    "agent_name": row["agent_name"],
                    "agent_email": row["agent_email"],
                    "agent_phone": row["agent_phone"],
                    "agent_address": row["agent_address"],
                    "agent_vat": row["agent_vat"],
                    "agent_contact_person": row["agent_contact_person"],
                    "group_name": row["group_name"],
                    "group_email": row["group_email"],
                    "group_address": row["group_address"],
                    "group_phone": row["group_phone"],
                    "group_vat": row["group_vat"],
                    "group_contact_person": row["group_contact_person"],
                    "status": row["status"],
                    "comments": row["comments"]
                    
                }

                serializer = DomainPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class DomainPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            'domain_name', 'extension', 'creation_date', 'deadline', 
                'owner_name', 'owner_email', 'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
                'agent_name', 'agent_email', 'agent_phone', 'agent_address', 'agent_vat', 'agent_contact_person', 
                'group_name', 'group_email', 'group_address', 'group_phone', 'group_vat', 'group_contact_person', 
                'status', 'comments'
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="patent_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="PatentPortfolio", index=False)

        return response
    


# Views in brand/trademark Portfolio 

class BrandPortfolioViewSet(viewsets.ModelViewSet):
    queryset = BrandPortfolio.objects.all().order_by('-updated_date')
    serializer_class = BrandPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  BrandPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ["title",  "territory", "b_Type" ,  "deadline", "status",
        "owner_name", "agent_name" , "group_name", "affidavit"]


    def list(self, request, *args, **kwargs):
        """Fetch all Brand records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(brand)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new brand record."""
        print( request.data , "post request")
        # Get the uploaded file
        # Check if logo is a string (existing URL) and remove it from the request data
        logo = request.data.get('logo')
        request_data = request.data.copy()
        print( logo , type(logo))
        if isinstance(logo, str):
            if logo is None : 
                print("User requested logo removal. Setting logo to None.")
                request_data['logo'] = None  # Mark logo for removal
            elif logo == 'http://127.0.0.1:8000/media/na.png':
                print("User requested logo removal. Setting logo to None.")
                request_data['logo'] = None  # Mark logo for removal
                # request_data['logo'] = os.path.join('logo', 'na.png') 
            else:
                print("Existing logo detected, skipping logo update.")
                request_data.pop('logo', None)  # Remove from request to prevent update
        print( request_data , "final data" )
        serializer = self.get_serializer( data=request_data)
        
        
        #serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print(request.data, "test update")

        # Extract logo from request
        logo = request.data.get('logo')
        print(f"Logo Type: {type(logo)}, Logo Value: {logo}")

        try:
            brand = BrandPortfolio.objects.get(pk=pk)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Trademark not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check if logo is a string (existing URL) and remove it from the request data
        request_data = request.data.copy()
        print( logo , type(logo))
        if logo is None : 
            print( "No image is uploaded")
            request_data['logo'] = None
        elif isinstance(logo, str):
            if logo == 'http://127.0.0.1:8000/media/na.png':
                print("User requested logo removal. Setting logo to None.")
                request_data['logo'] = None  # Mark logo for removal
                # request_data['logo'] = os.path.join('logo', 'na.png') 
            else:
                print("Existing logo detected, skipping logo update.")
                request_data.pop('logo', None)  # Remove from request to prevent update
        print( request_data , "final data" )
        serializer = self.get_serializer( brand, data=request_data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            print("no valid data")
            print("Validation errors:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( brand, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single brand record by ID."""
        try:
            brand = BrandPortfolio.objects.get(pk=pk)
            brand.delete()
            return Response({"message": "Brand deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except BrandPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple brand records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = BrandPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} brands deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = BrandPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            "Title", "Territory", "Type", "Application Date", "Application No", "Registration Date", "Registration No", 
            "Classes", "Deadline", "Affidavit", "Owner Name", "Owner Email", "Owner Phone", "Owner Address", 
            "Owner VAT", "Owner Contact Person", "Agent Name", "Agent Email", "Agent Phone", "Agent Address", 
            "Agent VAT", "Agent Contact Person", "Group Name", "Group Email", "Group Address", "Group Phone", 
            "Group VAT", "Group Contact Person", "Comments", "Status", "Updated Date", "Logo"
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.title, row.territory, row.b_Type, str(row.application_date), row.application_no,
                str(row.registration_date), row.registration_no, row.classes, str(row.deadline), str(row.affidavit),
                row.owner_name, row.owner_email, row.owner_phone, row.owner_address, row.owner_vat, row.owner_contact_person,
                row.agent_name, row.agent_email, row.agent_phone, row.agent_address, row.agent_vat, row.agent_contact_person,
                row.group_name, row.group_email, row.group_address, row.group_phone, row.group_vat, row.group_contact_person,
                row.comments, row.status, str(row.updated_date)
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else 'N/A', text_wrap_format)

            # Handle Image Download & Insertion with formatting
            if row.logo and row.logo != 'na.png':
                logo_url = request.build_absolute_uri(f'{settings.MEDIA_URL}{row.logo}')
                try:
                    image_response = requests.get(logo_url, stream=True)
                    if image_response.status_code == 200:
                        image_path = f"/tmp/{row.id}.png"
                        with open(image_path, "wb") as img_file:
                            img_file.write(image_response.content)

                        from PIL import Image
                        img = Image.open(image_path)
                        width, height = img.size
                        scale_x = 100 / width
                        scale_y = 100 / height
                        scale = min(scale_x, scale_y)
                        worksheet.set_row(row_num, 100)  # Set row height to a fixed size
                        # Write a blank cell with border format before inserting the image
                        worksheet.write_blank(row_num, len(headers) - 1, '', border_format)

                        # Insert image after ensuring the cell has a border
                        worksheet.insert_image(row_num, len(headers) - 1, image_path, {'x_scale': scale, 'y_scale': scale})
                    else:
                        worksheet.write(row_num, len(headers) - 1, "Image not available", border_format)
                except Exception as e:
                    print(f"Failed to download image for {row.title}: {e}")
                    worksheet.write(row_num, len(headers) - 1, "Error loading image", border_format)
            else:
                worksheet.write(row_num, len(headers) - 1, "No Image", border_format)

            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_brands.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple brand records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for brand in BrandPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(brand, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr(brand, key, value)

            # Check and update logo only if provided
            print("before" , logo)
            if logo is not None:
                print("after" , logo)
                current_logo = brand.logo.url if brand.logo else None
                changes['logo'] = {"old": current_logo, "new": logo.name}
                brand.logo = logo

            if changes:
                print(f"Updating Brand ID {brand.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            brand_data_before_save = model_to_dict(brand)
            print(f"Brand Data Before Save (ID: {brand.id}):", brand_data_before_save)
            brand.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} brands updated successfully"}, status=status.HTTP_200_OK)



class UploadBrandXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            image_loader = SheetImageLoader(sheet)  # Load images

            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            # Convert to datetime if it's not already
            df["application_date"] = pd.to_datetime(df["application_date"], errors='coerce')
            df["registration_date"] = pd.to_datetime(df["registration_date"], errors='coerce')
            df["deadline"] = pd.to_datetime(df["deadline"], errors='coerce')
            df["affidavit"] = pd.to_datetime(df["affidavit"], errors='coerce')


            # Handle invalid dates (NaT) by replacing with None or a default date if needed
            df["application_date"] = df["application_date"].dt.strftime("%Y-%m-%d").fillna("")
            df["registration_date"] = df["registration_date"].dt.strftime("%Y-%m-%d").fillna("")
            df["deadline"] = df["deadline"].dt.strftime("%Y-%m-%d").fillna("")
            df["affidavit"] = df["affidavit"].dt.strftime("%Y-%m-%d").fillna("")
            
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                "title", "logo", "b_type", "territory", "application_date", "application_no",
                "registration_date", "registration_no", "classes", "deadline", "affidavit",
                "owner_name", "owner_email", "owner_phone", "owner_address", "owner_vat",
                "owner_contact_person", "agent_name", "agent_email", "agent_phone",
                "agent_address", "agent_vat", "agent_contact_person", "group_name",
                "group_email", "group_address", "group_phone", "group_vat",
                "group_contact_person", "status", "comments"
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Get today's date for folder structuring
            today = datetime.today()
            year, month, day = today.year, today.month, today.day
            image_folder = Path(settings.MEDIA_ROOT) / "logos" / str(year) / str(month) / str(day)
            image_folder.mkdir(parents=True, exist_ok=True)  # Create if not exists

            # Process rows
            records = []
            max_id = BrandPortfolio.objects.all().count()  # Get max ID for unique naming

            print("Starting iterations")
            for index, row in df.iterrows():
                print( index , row )
                logo_image_path = None

                # Extract and save image if present
                image_cell = f"B{index + 2}"  # Assuming logos are in column B
                if image_loader.image_in(image_cell):
                    print("Saving image from cell")
                    img = image_loader.get(image_cell)
        
                    max_id += 1
                    image_name = f"image_{max_id}.png"
                    image_path = image_folder / image_name
                    img.save(image_path)
                    logo_image_path = f"logos/{year}/{month}/{day}/{image_name}"
                    print("Image loaded")

                    # Convert the image to an in-memory file for posting
                    image_io = BytesIO()
                    img.save(image_io, format="PNG")
                    image_io.seek(0)  # Reset pointer to start

                    # Convert BytesIO to InMemoryUploadedFile
                    logo_image_file = InMemoryUploadedFile(
                        file=image_io,  # The in-memory file
                        field_name=None,
                        name=image_name,  # File name
                        content_type="image/png",
                        size=sys.getsizeof(image_io),
                        charset=None
                    )

                # Prepare data for saving
                # Format Date type fields \
                print( type( row["application_date"]) , type(logo_image_file) , "type check")

                data = {
                    "title": row["title"],
                    "logo": logo_image_file ,  
                    "b_Type": row["b_type"],
                    "territory": row["territory"],
                    "application_date": row["application_date"],
                    "application_no": row["application_no"],
                    "registration_date": row["registration_date"],
                    "registration_no": row["registration_no"],
                    "classes": row["classes"],
                    "deadline": row["deadline"],
                    "affidavit": row["affidavit"],
                    "owner_name": row["owner_name"],
                    "owner_email": row["owner_email"],
                    "owner_phone": row["owner_phone"],
                    "owner_address": row["owner_address"],
                    "owner_vat": row["owner_vat"],
                    "owner_contact_person": row["owner_contact_person"],
                    "agent_name": row["agent_name"],
                    "agent_email": row["agent_email"],
                    "agent_phone": row["agent_phone"],
                    "agent_address": row["agent_address"],
                    "agent_vat": row["agent_vat"],
                    "agent_contact_person": row["agent_contact_person"],
                    "group_name": row["group_name"],
                    "group_email": row["group_email"],
                    "group_address": row["group_address"],
                    "group_phone": row["group_phone"],
                    "group_vat": row["group_vat"],
                    "group_contact_person": row["group_contact_person"],
                    "status": row["status"],
                    "comments": row["comments"],
                }

                serializer = BrandPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class BrandPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            "title", 
                "logo",
                "b_Type", 
                "territory",
                "application_date",
                "application_no",
                "registration_date",
                "registration_no",
                "classes",
                "deadline",
                "affidavit",
                "owner_name", 
                "owner_email", 
                "owner_phone", 
                "owner_address", 
                "owner_vat", 
                "owner_contact_person", 
                "agent_name", 
                "agent_email" , 
                "agent_phone", 
                "agent_address", 
                "agent_vat", 
                "agent_contact_person",
                "group_name",
                "group_email",
                "group_address",
                "group_phone",
                "group_vat",
                "group_contact_person",
                "status",
                "comments"
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="patent_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="PatentPortfolio", index=False)

        return response
    


# Views in Model/Design Portfolio 

class ModelPortfolioViewSet(viewsets.ModelViewSet):
    queryset = ModelPortfolio.objects.all().order_by('-updated_date')
    serializer_class = ModelPortfolioSerializer
    permission_classes = [IsAuthenticated]  # Protects API with authentication
    filter_backends = [DjangoFilterBackend, SearchFilter]
    
    filterset_class =  ModelPortfolioFilter

    # Searching in important text fields (avoiding file fields)
    search_fields = ["design_title",  "territory", "d_Type" ,  "deadline", "status",
        "owner_name", "agent_name" , "group_name"]


    def list(self, request, *args, **kwargs):
        """Fetch all Brand records with search and filtering capabilities."""
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None, *args, **kwargs):
        """Fetch a single Brand record by ID."""
        try:
            brand = ModelPortfolio.objects.get(pk=pk)
            serializer = self.get_serializer(brand)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ModelPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        """Create a new brand record."""
        print( request.data , "post request add")
        logo = request.data.get('logo')
        # Get the uploaded file
        # Check if logo is a string (existing URL) and remove it from the request data
        request_data = request.data.copy()
        print( logo , type(logo))
        if isinstance(logo, str):
            if logo == 'http://127.0.0.1:8000/media/na.png':
                print("User requested logo removal. Setting logo to None.")
                request_data['logo'] = None  # Mark logo for removal
                # request_data['logo'] = os.path.join('logo', 'na.png') 
            else:
                print("Existing logo detected, skipping logo update.")
                request_data.pop('logo', None)  # Remove from request to prevent update
        print( request_data , "final data" )
        serializer = self.get_serializer(data=request_data)
        # serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else : 
            print("no valid data")
            print("Validation errors:", serializer.errors) 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a single patent record by ID."""
        print(request.data, "test update")

        # Extract logo from request
        logo = request.data.get('logo')
        print(f"Logo Type: {type(logo)}, Logo Value: {logo}")

        try:
            model = ModelPortfolio.objects.get(pk=pk)
        except ModelPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check if logo is a string (existing URL) and remove it from the request data
        request_data = request.data.copy()
        print( logo , type(logo))
        if isinstance(logo, str):
            if logo == 'http://127.0.0.1:8000/media/na.png':
                print("User requested logo removal. Setting logo to None.")
                request_data['logo'] = None  # Mark logo for removal
                # request_data['logo'] = os.path.join('logo', 'na.png') 
            else:
                print("Existing logo detected, skipping logo update.")
                request_data.pop('logo', None)  # Remove from request to prevent update
        print( request_data , "final data" )
        serializer = self.get_serializer(model, data=request_data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            print("no valid data")
            print("Validation errors:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None, *args, **kwargs):
        """Partially update a single brand record by ID."""
        try:
            model = ModelPortfolio.objects.get(pk=pk)
        except ModelPortfolio.DoesNotExist:
            return Response({"error": "brand not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer( model , data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None, *args, **kwargs):
        """Delete a single brand record by ID."""
        try:
            model = ModelPortfolio.objects.get(pk=pk)
            model.delete()
            return Response({"message": "Brand deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except ModelPortfolio.DoesNotExist:
            return Response({"error": "Brand not found"}, status=status.HTTP_404_NOT_FOUND)
        
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """
        Delete multiple brand records by IDs.
        Example Payload: {"ids": [1, 2, 3]}
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Perform deletion
        deleted_count, _ = ModelPortfolio.objects.filter(id__in=ids).delete()
        
        if deleted_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({"message": f"{deleted_count} brands deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
    # Api to fetch selected as download 
    @action(detail=False, methods=['post'])
    def download_selected(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch selected rows
        selected_rows = ModelPortfolio.objects.filter(id__in=ids)
        print(selected_rows.values(), "selected_rows")

        # Create an in-memory Excel file
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Define column headers
        headers = [
            'design_title', 'logo', 'territory', 'd_Type', 
            'application_date', 'application_no', 'registration_date', 
            'registration_no', 'deadline', 'owner_name', 'owner_email', 
            'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
            'agent_name', 'agent_email', 'agent_phone', 'agent_address', 'agent_vat', 
            'agent_contact_person', 'group_name', 'group_email', 'group_address', 
            'group_phone', 'group_vat', 'group_contact_person', 'inventor', 'comments', 
            'status'
        ]

        # Formatting
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
        border_format = workbook.add_format({'border': 1})
        text_wrap_format = workbook.add_format({'text_wrap': True, 'border': 1})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Adjust column widths
        for col_num in range(len(headers) - 1):
            worksheet.set_column(col_num, col_num, 20)
        worksheet.set_column(len(headers) - 1, len(headers) - 1, 30)

        # Insert data into Excel
        row_num = 1
        for row in selected_rows:
            data = [
                row.design_title, row.logo, row.territory, row.d_Type,  
                str(row.application_date), row.application_no, row.registration_date,  
                row.registration_no, row.deadline, row.owner_name, row.owner_email, 
                row.owner_phone, row.owner_address, row.owner_vat, row.owner_contact_person, 
                row.agent_name, row.agent_email, row.agent_phone, row.agent_address, row.agent_vat, 
                row.agent_contact_person, row.group_name, row.group_email, row.group_address, 
                row.group_phone, row.group_vat, row.group_contact_person, row.inventor, row.comments,  
                row.status
            ]

            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value if value else '', text_wrap_format)

            # Handle Image Download & Insertion with formatting
            if row.logo and row.logo != 'na.png':
                logo_url = request.build_absolute_uri(f'{settings.MEDIA_URL}{row.logo}')
                try:
                    image_response = requests.get(logo_url, stream=True)
                    if image_response.status_code == 200:
                        image_path = f"/tmp/{row.id}.png"
                        with open(image_path, "wb") as img_file:
                            img_file.write(image_response.content)

                        from PIL import Image
                        img = Image.open(image_path)
                        width, height = img.size
                        scale_x = 100 / width
                        scale_y = 100 / height
                        scale = min(scale_x, scale_y)
                        worksheet.set_row(row_num, 100)  # Set row height to a fixed size
                        # Write a blank cell with border format before inserting the image
                        worksheet.write_blank(row_num, len(headers) - 1, '', border_format)

                        # Insert image after ensuring the cell has a border
                        worksheet.insert_image(row_num, len(headers) - 1, image_path, {'x_scale': scale, 'y_scale': scale})
                    else:
                        worksheet.write(row_num, len(headers) - 1, "Image not available", border_format)
                except Exception as e:
                    print(f"Failed to download image for {row.title}: {e}")
                    worksheet.write(row_num, len(headers) - 1, "Error loading image", border_format)
            else:
                worksheet.write(row_num, len(headers) - 1, "No Image", border_format)

            row_num += 1

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=selected_brands.xlsx'
        return response
    
    @action(detail=False, methods=['post'], url_path='batch-update')
    def batch_update(self, request):
        """
        Update multiple brand records by IDs including logo.
        """
        print("Received Data:", request.data)
        print("Received Files:", request.FILES)

        # Extract IDs and data
        ids = request.data.getlist('ids[]', [])
        data = {
            key.replace('data[', '').replace(']', ''): value
            for key, value in request.data.items() if key.startswith('data[')
        }
        logo = request.FILES.get('logo')

        print(ids, data, logo, "input data")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        if not data and not logo:
            return Response({"error": "No update data provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Perform updates
        updated_count = 0
        for model in ModelPortfolio.objects.filter(id__in=ids):
            changes = {}

            # Check and update non-logo fields
            print( data.items() , "Data Items" )
            for key, value in data.items():
                if key != 'logoPreview' and value is not None:
                    if key != 'logo' and value != 'na.png':
                        current_value = getattr(model, key, None)
                        if str(current_value) != str(value):
                            changes[key] = {"old": current_value, "new": value}
                            setattr( model, key, value)

            # Check and update logo only if provided
            print("before" , logo)
            if logo is not None:
                print("after" , logo)
                current_logo = model.logo.url if model.logo else None
                changes['logo'] = {"old": current_logo, "new": logo.name}
                model.logo = logo

            if changes:
                print(f"Updating Brand ID {model.id} with changes: {changes}")

             # Print the updated values using model_to_dict
            from django.forms.models import model_to_dict
            model_data_before_save = model_to_dict( model)
            print(f"Brand Data Before Save (ID: { model.id}):", model_data_before_save)
            model.save()
            updated_count += 1

        if updated_count == 0:
            return Response({"error": "No matching records found"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"message": f"{updated_count} brands updated successfully"}, status=status.HTTP_200_OK)



class UploadModelXLSViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Secure API

    def create(self, request, *args, **kwargs):
        print("Starting upload", settings.MEDIA_ROOT)
        
        file_obj = request.FILES.get("file")  # Get uploaded file
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the directory exists
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save the file temporarily
        upload_path = os.path.join("uploads", file_obj.name)
        saved_path = default_storage.save(upload_path, ContentFile(file_obj.read()))
        file_path = default_storage.path(saved_path)  # Get absolute path

        print(f"File saved at {file_path}")

        #try:
        if 1 ==1 : 
            # Load Excel file
            wb = load_workbook(file_path)
            sheet = wb.active  # Use first sheet
            image_loader = SheetImageLoader(sheet)  # Load images

            # Read data using pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            df["application_date"] = df["application_date"].dt.strftime("%Y-%m-%d")
            df["registration_date"] = df["registration_date"].dt.strftime("%Y-%m-%d")
            df["deadline"] = df["deadline"].dt.strftime("%Y-%m-%d")
            # df["affidavit"] = df["affidavit"].dt.strftime("%Y-%m-%d")
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            print(f"Renamed Columns: {df.columns.tolist()}")

            # Validate required columns
            required_columns = [
                'design_title', 'logo', 'territory', 'd_type', 
                'application_date', 'application_no', 'registration_date', 
                'registration_no', 'deadline', 'owner_name', 'owner_email', 
                'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
                'agent_name', 'agent_email', 'agent_phone', 'agent_address', 
                'agent_vat', 'agent_contact_person', 'group_name', 'group_email', 
                'group_address', 'group_phone', 'group_vat', 'group_contact_person', 
                'inventor', 'comments', 'status'
            ]
            missing_columns = set(required_columns) - set(df.columns)
            if missing_columns:
                return Response({"error": f"Missing required columns: {', '.join(missing_columns)}"}, status=status.HTTP_400_BAD_REQUEST)

            # Get today's date for folder structuring
            today = datetime.today()
            year, month, day = today.year, today.month, today.day
            image_folder = Path(settings.MEDIA_ROOT) / "logos" / str(year) / str(month) / str(day)
            image_folder.mkdir(parents=True, exist_ok=True)  # Create if not exists

            # Process rows
            records = []
            print(" getting max id ")
            max_id = ModelPortfolio.objects.all().count()  # Get max ID for unique naming
            print( max_id , "max id ")
            print("Starting iterations")
            for img in sheet._images:
                print(f"Image at cell: {img.anchor._from.row + 1}, {img.anchor._from.col + 1}")
            for index, row in df.iterrows():
                print( index , row )
                logo_image_path = None

                # Extract and save image if present
                image_cell = f"B{index + 2}"  # Assuming logos are in column B
                if image_loader.image_in(image_cell):
                    print("Saving image from cell")
                    img = image_loader.get(image_cell)
        
                    max_id += 1
                    image_name = f"image_{max_id}.png"
                    image_path = image_folder / image_name
                    img.save(image_path)
                    logo_image_path = f"logos/{year}/{month}/{day}/{image_name}"
                    print("Image loaded")

                    # Convert the image to an in-memory file for posting
                    image_io = BytesIO()
                    img.save(image_io, format="PNG")
                    image_io.seek(0)  # Reset pointer to start

                    # Convert BytesIO to InMemoryUploadedFile
                    logo_image_file = InMemoryUploadedFile(
                        file=image_io,  # The in-memory file
                        field_name=None,
                        name=image_name,  # File name
                        content_type="image/png",
                        size=sys.getsizeof(image_io),
                        charset=None
                    )

                # Prepare data for saving
                # Format Date type fields \
                #print( type( row["application_date"]) , type(logo_image_file) , "type check")

                data = {
                    'design_title' : row['design_title'], 
                    'logo' : logo_image_file , 
                    'territory' : row['territory'], 
                    'd_Type' : row['d_type'], 
                    'application_date' : row['application_date'], 
                    'application_no' : row['application_no'], 
                    'registration_date' : row['registration_date'],
                    'registration_no' : row['registration_no'], 
                    'deadline' : row['deadline'], 
                    'owner_name' : row['owner_name'], 
                    'owner_email' : row['owner_email'], 
                    'owner_phone' : row['owner_phone'], 
                    'owner_address' : row['owner_address'], 
                    'owner_vat' : row['owner_vat'], 
                    'owner_contact_person' : row["owner_contact_person"], 
                    'agent_name' : row["agent_name"], 
                    'agent_email' : row["agent_email"], 
                    'agent_phone' : row["agent_phone"], 
                    'agent_address' : row["agent_address"], 
                    'agent_vat' : row["agent_vat"], 
                    'agent_contact_person' : row["agent_contact_person"], 
                    'group_name' : row["group_name"], 
                    'group_email' : row["group_email"], 
                    'group_address' : row["group_address"], 
                    'group_phone' : row["agent_phone"], 
                    'group_vat' : row["group_vat"], 
                    'group_contact_person' : row["group_contact_person"], 
                    'inventor' : row["inventor"], 
                    'comments' : row["comments"], 
                    'status' : row["status"]
                    
                }

                serializer = ModelPortfolioSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    records.append(serializer.data)
                else:
                    print("no valid data")
                    print("Validation errors:", serializer.errors) 
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "File processed successfully", "data": records}, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     print( str(e))
        #     return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # finally:
        #     # Clean up temporary file
        #     if os.path.exists(file_path):
        #         os.remove(file_path)


class ModelPortfolioUploadViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]  # Protect API with authentication

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        # Define the columns expected in the template
        column_headers = [
            'design_title', 'logo', 'territory', 'd_Type', 
                'application_date', 'application_no', 'registration_date', 
                'registration_no', 'deadline', 'owner_name', 'owner_email', 
                'owner_phone', 'owner_address', 'owner_vat', 'owner_contact_person', 
                'agent_name', 'agent_email', 'agent_phone', 'agent_address', 
                'agent_vat', 'agent_contact_person', 'group_name', 'group_email', 
                'group_address', 'group_phone', 'group_vat', 'group_contact_person', 
                'inventor', 'comments', 'status'
        ]

        # Create an empty DataFrame with headers
        df = pd.DataFrame(columns=column_headers)

        # Save to an in-memory file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="design_portfolio_template.xlsx"'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name="DesignPortfolio", index=False)

        return response
    





    



        




    

